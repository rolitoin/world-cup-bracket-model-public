import copy
import random
import hashlib
import json
import math
from pathlib import Path

import openpyxl


INPUT = Path("sheet_download.xlsx")
OUTPUT = Path("world_cup_bracket_model_picks.xlsx")
SUMMARY = Path("world_cup_bracket_model_summary.json")
SCORE_CALIBRATION = Path("data/score_calibration.json")
FIFA_RANKINGS = Path("data/fifa_top20_rankings_2026_04_01.json")
RECENT_FORM = Path("data/recent_form_adjustments.json")
SQUAD_PROFILE = Path("data/squad_profile_adjustments.json")
VENUE_CONTEXT = Path("data/venue_context_2026.json")
COACH_ADJUSTMENTS = Path("data/coach_adjustments.json")
MONTE_CARLO_TRIALS = 5000

GROUP_ROWS = {
    "A": (6, 11),
    "B": (15, 20),
    "C": (24, 29),
    "D": (33, 38),
    "E": (42, 47),
    "F": (51, 56),
    "G": (60, 65),
    "H": (69, 74),
    "I": (78, 83),
    "J": (87, 92),
    "K": (96, 101),
    "L": (105, 110),
}

HOSTS = {"México": 4.5, "EE.UU.": 4.0, "Canadá": 3.0}

FORM_EXPERT = {
    "Francia": 3.0,
    "España": 2.5,
    "Inglaterra": 2.0,
    "Brasil": 2.0,
    "Argentina": 1.5,
    "Alemania": 1.5,
    "Colombia": 2.5,
    "Noruega": 2.0,
    "Japón": 1.5,
    "Marruecos": 1.5,
    "Suiza": 1.0,
    "Turquía": 1.0,
    "México": 1.0,
    "Uruguay": 1.0,
    "Portugal": 0.5,
    "Países Bajos": 0.5,
    "EE.UU.": 0.5,
}

HEAT_TRAVEL = {
    "México": 1.5,
    "EE.UU.": 1.5,
    "Canadá": 0.5,
    "Brasil": 1.0,
    "Colombia": 1.0,
    "Ecuador": 1.0,
    "Paraguay": 1.0,
    "Uruguay": 0.5,
    "Marruecos": 1.0,
    "Senegal": 1.0,
    "Ghana": 1.0,
    "Costa de Marfil": 1.0,
    "Egipto": 0.5,
    "Túnez": 0.5,
    "Catar": 0.5,
    "Arabia Saudita": 0.5,
    "Países Bajos": -0.5,
    "Suecia": -0.5,
    "Noruega": -0.5,
    "Inglaterra": -0.3,
    "Escocia": -0.3,
}

CLIMATE_ADAPTATION = {
    "México": 0.9,
    "EE.UU.": 0.4,
    "Canadá": -0.2,
    "Brasil": 0.8,
    "Colombia": 0.8,
    "Ecuador": 0.7,
    "Paraguay": 0.7,
    "Panamá": 0.8,
    "Uruguay": 0.3,
    "Argentina": 0.2,
    "Marruecos": 0.7,
    "Senegal": 0.8,
    "Ghana": 0.8,
    "Costa de Marfil": 0.8,
    "Sudáfrica": 0.4,
    "Egipto": 0.7,
    "Túnez": 0.6,
    "Argelia": 0.6,
    "Haití": 0.8,
    "Catar": 0.8,
    "Arabia Saudita": 0.9,
    "Irán": 0.6,
    "Irak": 0.8,
    "Jordania": 0.7,
    "Japón": 0.2,
    "Corea del Sur": 0.1,
    "Australia": 0.4,
    "Nueva Zelanda": -0.1,
    "España": 0.3,
    "Portugal": 0.3,
    "Francia": 0.0,
    "Italia": 0.2,
    "Alemania": -0.4,
    "Inglaterra": -0.5,
    "Escocia": -0.7,
    "Países Bajos": -0.5,
    "Bélgica": -0.4,
    "Suiza": -0.4,
    "Austria": -0.4,
    "Croacia": -0.2,
    "Bosnia": -0.2,
    "Chequia": -0.4,
    "Suecia": -0.7,
    "Noruega": -0.8,
    "Turquía": 0.4,
    "Uzbekistán": 0.3,
    "Cabo Verde": 0.7,
    "Curazao": 0.8,
    "RD Congo": 0.8,
}

ENERGY_EXPERIENCE = {
    "España": 1.5,
    "Inglaterra": 1.0,
    "Francia": 1.0,
    "Alemania": 0.8,
    "Japón": 1.0,
    "EE.UU.": 0.8,
    "Colombia": 0.8,
    "Marruecos": 0.8,
    "Austria": 0.7,
    "Noruega": 0.5,
    "Argentina": -0.5,
    "Portugal": -0.5,
    "Croacia": -1.0,
    "Bélgica": -0.5,
    "Bosnia": -0.6,
}

UPSET_PROFILE = {
    "Marruecos": 4.0,
    "Croacia": 3.5,
    "Japón": 3.5,
    "Colombia": 3.0,
    "Noruega": 3.0,
    "Turquía": 3.0,
    "Ecuador": 2.5,
    "Senegal": 2.5,
    "Austria": 2.5,
    "Suecia": 2.0,
    "Paraguay": 2.0,
    "Corea del Sur": 2.0,
    "Ghana": 2.0,
    "México": 1.5,
    "EE.UU.": 1.5,
    "Canadá": 1.5,
}


def stable_noise(*parts):
    text = "|".join(str(p) for p in parts)
    raw = hashlib.sha256(text.encode("utf-8")).hexdigest()[:8]
    return int(raw, 16) / 0xFFFFFFFF


def clamp(value, low, high):
    return max(low, min(high, value))


def poisson_pmf(goals, expected_goals):
    return (expected_goals**goals) * math.exp(-expected_goals) / math.factorial(goals)


def expected_goals(home, away, match_id, teams, context, knockout=False):
    home_rating = rating(home, teams, context, match_id=match_id, knockout=knockout)
    away_rating = rating(away, teams, context, match_id=match_id, knockout=knockout)
    diff = home_rating - away_rating
    combined_quality = (teams[home] + teams[away]) / 100
    total_goals = 2.35 + combined_quality * 0.35
    if knockout:
        total_goals -= 0.12
    if abs(diff) < 4:
        total_goals -= 0.12
    total_goals = clamp(total_goals, 1.85, 3.25)

    home_share = clamp(0.5 + diff / 72, 0.18, 0.82)
    home_xg = clamp(total_goals * home_share, 0.25, 3.2)
    away_xg = clamp(total_goals - home_xg, 0.25, 3.2)
    return home_xg, away_xg


def score_grid(home_xg, away_xg, max_goals=6):
    scores = []
    home_prob = draw_prob = away_prob = 0
    for home_goals in range(max_goals + 1):
        for away_goals in range(max_goals + 1):
            probability = poisson_pmf(home_goals, home_xg) * poisson_pmf(away_goals, away_xg)
            scores.append((probability, home_goals, away_goals))
            if home_goals > away_goals:
                home_prob += probability
            elif home_goals == away_goals:
                draw_prob += probability
            else:
                away_prob += probability

    total = home_prob + draw_prob + away_prob
    return {
        "scores": scores,
        "home": home_prob / total,
        "draw": draw_prob / total,
        "away": away_prob / total,
    }


def load_score_calibration():
    if not SCORE_CALIBRATION.exists():
        return {"winner_loser_patterns": {}, "draw_patterns": {}, "matches": 0}
    return json.loads(SCORE_CALIBRATION.read_text(encoding="utf-8"))


def load_model_context():
    fifa = {}
    if FIFA_RANKINGS.exists():
        data = json.loads(FIFA_RANKINGS.read_text(encoding="utf-8"))
        fifa_name_map = {
            "France": "Francia",
            "Spain": "España",
            "Argentina": "Argentina",
            "England": "Inglaterra",
            "Portugal": "Portugal",
            "Brazil": "Brasil",
            "Netherlands": "Países Bajos",
            "Morocco": "Marruecos",
            "Belgium": "Bélgica",
            "Germany": "Alemania",
            "Croatia": "Croacia",
            "Colombia": "Colombia",
            "Senegal": "Senegal",
            "Mexico": "México",
            "United States": "EE.UU.",
            "Uruguay": "Uruguay",
            "Japan": "Japón",
            "Switzerland": "Suiza",
        }
        for row in data.get("rankings", []):
            team = fifa_name_map.get(row["team"], row["team"])
            fifa[team] = row

    recent_form = {}
    if RECENT_FORM.exists():
        data = json.loads(RECENT_FORM.read_text(encoding="utf-8"))
        recent_form = {row["team"]: row for row in data.get("adjustments", [])}

    squad_profile = {}
    if SQUAD_PROFILE.exists():
        data = json.loads(SQUAD_PROFILE.read_text(encoding="utf-8"))
        squad_profile = {row["team"]: row for row in data.get("adjustments", [])}

    coaches = {}
    if COACH_ADJUSTMENTS.exists():
        data = json.loads(COACH_ADJUSTMENTS.read_text(encoding="utf-8"))
        coaches = {row["team"]: row for row in data.get("adjustments", [])}

    venues = {}
    if VENUE_CONTEXT.exists():
        data = json.loads(VENUE_CONTEXT.read_text(encoding="utf-8"))
        for row in data.get("fixtures", []):
            venues[row["local_match"]] = row
        for row in data.get("knockout_venues", []):
            venues[row["official_match"]] = row

    return {
        "fifa": fifa,
        "recent_form": recent_form,
        "squad_profile": squad_profile,
        "coaches": coaches,
        "venues": venues,
    }


def score_pattern(home_goals, away_goals):
    if home_goals == away_goals:
        return "draw", f"{home_goals}-{away_goals}"
    return "win", f"{max(home_goals, away_goals)}-{min(home_goals, away_goals)}"


def score_prior_multiplier(home_goals, away_goals, calibration):
    kind, pattern = score_pattern(home_goals, away_goals)
    patterns = calibration.get("draw_patterns", {}) if kind == "draw" else calibration.get("winner_loser_patterns", {})
    if not patterns:
        return 1
    max_count = max(patterns.values())
    count = patterns.get(pattern, 0)
    return 0.35 + (count / max_count)


def sample_score(grid, rng, allow_draw=True, calibration=None):
    candidates = []
    total = 0
    for probability, home_goals, away_goals in grid["scores"]:
        if allow_draw or home_goals != away_goals:
            adjusted_probability = probability * score_prior_multiplier(home_goals, away_goals, calibration or {})
            candidates.append((adjusted_probability, home_goals, away_goals))
            total += adjusted_probability

    pick = rng.random() * total
    running = 0
    for probability, home_goals, away_goals in candidates:
        running += probability
        if running >= pick:
            return home_goals, away_goals
    return candidates[-1][1], candidates[-1][2]


def most_likely_score(grid, outcome, noise_key, calibration=None):
    candidates = []
    for probability, home_goals, away_goals in grid["scores"]:
        adjusted_probability = probability * score_prior_multiplier(home_goals, away_goals, calibration or {})
        if outcome == "home" and home_goals > away_goals:
            candidates.append((adjusted_probability, home_goals, away_goals))
        elif outcome == "draw" and home_goals == away_goals:
            candidates.append((adjusted_probability, home_goals, away_goals))
        elif outcome == "away" and away_goals > home_goals:
            candidates.append((adjusted_probability, home_goals, away_goals))

    candidates.sort(reverse=True)
    if not candidates:
        return 1, 1

    # Take the best score most of the time, but occasionally choose the second
    # plausible score so similar matchups do not all collapse into 1-0 or 2-1.
    roll = stable_noise("score", *noise_key)
    index = 0
    if len(candidates) > 1 and roll > 0.72 and candidates[1][0] >= candidates[0][0] * 0.72:
        index = 1
    if len(candidates) > 2 and roll > 0.9 and candidates[2][0] >= candidates[0][0] * 0.58:
        index = 2
    _, home_goals, away_goals = candidates[index]
    return home_goals, away_goals


def read_data(wb):
    ws = wb["Grupos"]
    teams = {}
    team_group = {}
    fixtures = []

    for group, (start, end) in GROUP_ROWS.items():
        for row in range(start, start + 4):
            team = ws[f"H{row}"].value
            strength = float(ws[f"N{row}"].value)
            teams[team] = strength
            team_group[team] = group

        for row in range(start, end + 1):
            fixtures.append(
                {
                    "match": int(ws[f"A{row}"].value),
                    "row": row,
                    "group": group,
                    "date": ws[f"B{row}"].value,
                    "home": ws[f"C{row}"].value,
                    "away": ws[f"F{row}"].value,
                }
            )

    return teams, team_group, fixtures


def fifa_adjustment(team, context):
    row = context.get("fifa", {}).get(team)
    if not row:
        return 0
    rank = row["rank"]
    return round((21 - rank) / 20 * 2.2, 2)


def recent_form_adjustment(team, context):
    row = context.get("recent_form", {}).get(team)
    if not row:
        return 0
    return row.get("adjustment", 0)


def squad_profile_adjustment(team, context):
    row = context.get("squad_profile", {}).get(team)
    if not row:
        return 0
    return row.get("adjustment", 0)


def coach_adjustment(team, context):
    row = context.get("coaches", {}).get(team)
    if not row:
        return 0
    return row.get("adjustment", 0)


def venue_context_adjustment(team, context, match_id):
    venue = context.get("venues", {}).get(match_id)
    if not venue:
        return 0
    climate = venue.get("stress", 0) * CLIMATE_ADAPTATION.get(team, 0) * 0.75
    travel = venue.get("travel_adjustments", {}).get(team, 0)
    return round(clamp(climate + travel, -1.2, 1.2), 2)


def rating(team, teams, context, match_id=None, knockout=False):
    value = teams[team]
    value += fifa_adjustment(team, context)
    value += recent_form_adjustment(team, context)
    value += squad_profile_adjustment(team, context)
    value += coach_adjustment(team, context)
    value += venue_context_adjustment(team, context, match_id)
    value += HOSTS.get(team, 0)
    value += FORM_EXPERT.get(team, 0)
    value += HEAT_TRAVEL.get(team, 0)
    value += ENERGY_EXPERIENCE.get(team, 0)
    if knockout and team in {"Argentina", "Portugal", "Croacia", "Francia", "Brasil", "Alemania"}:
        value += 0.8
    return value


def analyze_match(home, away, match_id, teams, context, knockout=False, calibration=None):
    home_rating = rating(home, teams, context, match_id=match_id, knockout=knockout)
    away_rating = rating(away, teams, context, match_id=match_id, knockout=knockout)
    diff = home_rating - away_rating

    underdog = home if diff < 0 else away
    upset_edge = UPSET_PROFILE.get(underdog, 0) - (abs(diff) / 2.8)
    upset_roll = stable_noise(match_id, home, away)
    force_upset = abs(diff) <= 8 and upset_edge > 0.6 and upset_roll > 0.62
    force_draw = not knockout and abs(diff) <= 4 and upset_roll < 0.38

    home_xg, away_xg = expected_goals(home, away, match_id, teams, context, knockout=knockout)
    grid = score_grid(home_xg, away_xg)

    if force_draw:
        outcome = "draw"
        reason = "draw-close"

    elif force_upset:
        outcome = "home" if underdog == home else "away"
        reason = f"upset-{underdog}"

    elif abs(diff) <= 2.5:
        if knockout:
            outcome = "draw"
            reason = "pens"
        else:
            outcome = "draw"
            reason = "draw-balance"

    else:
        if diff > 16:
            outcome = "home"
            reason = "favorite-clear"
        elif diff > 8:
            outcome = "home"
            reason = "favorite-solid"
        elif diff > 0:
            outcome = "home"
            reason = "favorite-narrow"
        elif diff < -16:
            outcome = "away"
            reason = "favorite-clear"
        elif diff < -8:
            outcome = "away"
            reason = "favorite-solid"
        else:
            outcome = "away"
            reason = "favorite-narrow"

    home_goals, away_goals = most_likely_score(grid, outcome, (match_id, home, away), calibration=calibration)
    if knockout and outcome == "draw" and home_goals == 0 and away_goals == 0:
        home_goals, away_goals = 1, 1
    if outcome == "home" and home_xg >= 1.75 and home_goals < 2:
        high_ceiling = home_xg >= 2.00 and diff >= 18 and stable_noise("margin", match_id, home, away) > 0.35
        home_goals = 3 if high_ceiling else 2
        away_goals = 1 if away_xg >= 0.50 and stable_noise("concede", match_id, home, away) > 0.55 else 0
    elif outcome == "away" and away_xg >= 1.75 and away_goals < 2:
        high_ceiling = away_xg >= 2.00 and diff <= -18 and stable_noise("margin", match_id, home, away) > 0.35
        away_goals = 3 if high_ceiling else 2
        home_goals = 1 if home_xg >= 0.50 and stable_noise("concede", match_id, home, away) > 0.55 else 0
    return {
        "home_goals": home_goals,
        "away_goals": away_goals,
        "reason": reason,
        "rating_diff": round(diff, 1),
        "venue_context": context.get("venues", {}).get(match_id, {}),
        "venue_adjustments": {
            home: venue_context_adjustment(home, context, match_id),
            away: venue_context_adjustment(away, context, match_id),
        },
        "expected_goals": {"home": round(home_xg, 2), "away": round(away_xg, 2)},
        "probabilities": {
            "home": round(grid["home"], 3),
            "draw": round(grid["draw"], 3),
            "away": round(grid["away"], 3),
        },
    }


def pick_score(home, away, match_id, teams, context, knockout=False, calibration=None):
    analysis = analyze_match(home, away, match_id, teams, context, knockout=knockout, calibration=calibration)
    return analysis["home_goals"], analysis["away_goals"], analysis["reason"]


def add_match(table, team, gf, ga, teams):
    table.setdefault(team, {"team": team, "pts": 0, "gf": 0, "ga": 0, "gd": 0, "strength": teams[team]})
    table[team]["gf"] += gf
    table[team]["ga"] += ga
    table[team]["gd"] = table[team]["gf"] - table[team]["ga"]
    if gf > ga:
        table[team]["pts"] += 3
    elif gf == ga:
        table[team]["pts"] += 1


def ordered_group_table(table):
    return sorted(table.values(), key=lambda x: (x["pts"], x["gd"], x["gf"], x["strength"]), reverse=True)


def group_results(fixtures, teams, context, calibration):
    scores = {}
    tables = {}
    rationale = []
    for fx in fixtures:
        analysis = analyze_match(fx["home"], fx["away"], fx["match"], teams, context, calibration=calibration)
        hg = analysis["home_goals"]
        ag = analysis["away_goals"]
        reason = analysis["reason"]
        scores[fx["match"]] = {**fx, "home_goals": hg, "away_goals": ag, "reason": reason}
        table = tables.setdefault(fx["group"], {})
        add_match(table, fx["home"], hg, ag, teams)
        add_match(table, fx["away"], ag, hg, teams)
        rationale.append(
            {
                "match": fx["match"],
                "group": fx["group"],
                "pick": f"{fx['home']} {hg}-{ag} {fx['away']}",
                "reason": reason,
                "rating_diff": analysis["rating_diff"],
                "venue": analysis["venue_context"].get("venue"),
                "venue_climate": analysis["venue_context"].get("climate"),
                "venue_adjustments": analysis["venue_adjustments"],
                "expected_goals": analysis["expected_goals"],
                "probabilities": analysis["probabilities"],
            }
        )

    standings = {}
    for group, rows in tables.items():
        ordered = ordered_group_table(rows)
        standings[group] = ordered

    thirds = []
    for group, ordered in standings.items():
        third = copy.deepcopy(ordered[2])
        third["group"] = group
        thirds.append(third)
    best_thirds = sorted(thirds, key=lambda x: (x["pts"], x["gd"], x["gf"], x["strength"]), reverse=True)[:8]
    third_groups = "".join(sorted(t["group"] for t in best_thirds))

    return scores, standings, best_thirds, third_groups, rationale


def simulated_group_results(fixtures, teams, context, rng, calibration):
    tables = {}
    for fx in fixtures:
        home_xg, away_xg = expected_goals(fx["home"], fx["away"], fx["match"], teams, context)
        grid = score_grid(home_xg, away_xg)
        hg, ag = sample_score(grid, rng, allow_draw=True, calibration=calibration)
        table = tables.setdefault(fx["group"], {})
        add_match(table, fx["home"], hg, ag, teams)
        add_match(table, fx["away"], ag, hg, teams)

    standings = {}
    for group, rows in tables.items():
        standings[group] = ordered_group_table(rows)

    thirds = []
    for group, ordered in standings.items():
        third = copy.deepcopy(ordered[2])
        third["group"] = group
        thirds.append(third)
    best_thirds = sorted(thirds, key=lambda x: (x["pts"], x["gd"], x["gf"], x["strength"]), reverse=True)[:8]
    third_groups = "".join(sorted(t["group"] for t in best_thirds))
    return standings, best_thirds, third_groups


def third_mapping(wb, third_groups):
    ws = wb["AnexoC"]
    for row in range(1, ws.max_row + 1):
        if ws[f"A{row}"].value == third_groups:
            return [ws.cell(row, col).value for col in range(2, 10)]
    raise ValueError(f"Missing AnexoC combination: {third_groups}")


def build_round_of_32(wb, standings, best_thirds, third_groups):
    third_by_group = {t["group"]: t["team"] for t in best_thirds}
    mapping = third_mapping(wb, third_groups)
    first = {g: standings[g][0]["team"] for g in standings}
    second = {g: standings[g][1]["team"] for g in standings}

    return [
        ("M73", second["A"], second["B"]),
        ("M74", first["E"], third_by_group[mapping[3]]),
        ("M75", first["F"], second["C"]),
        ("M76", first["C"], second["F"]),
        ("M77", first["I"], third_by_group[mapping[5]]),
        ("M78", second["E"], second["I"]),
        ("M79", first["A"], third_by_group[mapping[0]]),
        ("M80", first["L"], third_by_group[mapping[7]]),
        ("M81", first["D"], third_by_group[mapping[2]]),
        ("M82", first["G"], third_by_group[mapping[4]]),
        ("M83", second["K"], second["L"]),
        ("M84", first["H"], second["J"]),
        ("M85", first["B"], third_by_group[mapping[1]]),
        ("M86", first["J"], second["H"]),
        ("M87", first["K"], third_by_group[mapping[6]]),
        ("M88", second["D"], second["G"]),
    ]


def play_knockout_match(match_id, team1, team2, teams, context, calibration):
    analysis = analyze_match(team1, team2, match_id, teams, context, knockout=True, calibration=calibration)
    g1 = analysis["home_goals"]
    g2 = analysis["away_goals"]
    reason = analysis["reason"]
    pen = None
    if g1 == g2:
        diff = rating(team1, teams, context, match_id=match_id, knockout=True) - rating(
            team2, teams, context, match_id=match_id, knockout=True
        )
        if abs(diff) <= 5:
            upset1 = UPSET_PROFILE.get(team1, 0)
            upset2 = UPSET_PROFILE.get(team2, 0)
            winner = team1 if (diff + (upset1 - upset2) * 0.5) >= 0 else team2
        else:
            winner = team1 if diff > 0 else team2
        pen = 1 if winner == team1 else 2
    else:
        winner = team1 if g1 > g2 else team2
    return {
        "id": match_id,
        "team1": team1,
        "team2": team2,
        "g1": g1,
        "g2": g2,
        "pen": pen,
        "winner": winner,
        "reason": reason,
        "expected_goals": analysis["expected_goals"],
        "probabilities": analysis["probabilities"],
    }


def simulate_knockout_match(match_id, team1, team2, teams, context, rng, calibration):
    home_xg, away_xg = expected_goals(team1, team2, match_id, teams, context, knockout=True)
    grid = score_grid(home_xg, away_xg)
    g1, g2 = sample_score(grid, rng, allow_draw=False, calibration=calibration)
    return team1 if g1 > g2 else team2


def knockout(round32, teams, context, calibration):
    results = []
    current = []
    for i, (mid, team1, team2) in enumerate(round32, start=73):
        res = play_knockout_match(i, team1, team2, teams, context, calibration)
        results.append(res)
        current.append(res["winner"])

    # Mirror the spreadsheet formulas exactly.
    octavos_pairs = [(1, 4), (0, 2), (3, 5), (6, 7), (10, 11), (8, 9), (13, 15), (12, 14)]
    octavos = []
    for idx, (a, b) in enumerate(octavos_pairs, start=89):
        res = play_knockout_match(idx, current[a], current[b], teams, context, calibration)
        results.append(res)
        octavos.append(res["winner"])

    quarter_pairs = [(0, 1), (4, 5), (2, 3), (6, 7)]
    quarters = []
    for idx, (a, b) in enumerate(quarter_pairs, start=97):
        res = play_knockout_match(idx, octavos[a], octavos[b], teams, context, calibration)
        results.append(res)
        quarters.append(res["winner"])

    semi_pairs = [(0, 1), (2, 3)]
    semis = []
    semi_losers = []
    for idx, (a, b) in enumerate(semi_pairs, start=101):
        res = play_knockout_match(idx, quarters[a], quarters[b], teams, context, calibration)
        results.append(res)
        semis.append(res["winner"])
        semi_losers.append(res["team2"] if res["winner"] == res["team1"] else res["team1"])

    third = play_knockout_match(103, semi_losers[0], semi_losers[1], teams, context, calibration)
    final = play_knockout_match(104, semis[0], semis[1], teams, context, calibration)
    results.append(third)
    results.append(final)
    return results


def simulate_knockout(round32, teams, context, rng, calibration):
    current = []
    stage_counts = {
        "round_of_32": set(),
        "round_of_16": set(),
        "quarterfinal": set(),
        "semifinal": set(),
        "final": set(),
        "champion": set(),
    }

    for i, (_, team1, team2) in enumerate(round32, start=73):
        stage_counts["round_of_32"].update([team1, team2])
        current.append(simulate_knockout_match(i, team1, team2, teams, context, rng, calibration))

    octavos_pairs = [(1, 4), (0, 2), (3, 5), (6, 7), (10, 11), (8, 9), (13, 15), (12, 14)]
    octavos = []
    for idx, (a, b) in enumerate(octavos_pairs, start=89):
        team1 = current[a]
        team2 = current[b]
        stage_counts["round_of_16"].update([team1, team2])
        octavos.append(simulate_knockout_match(idx, team1, team2, teams, context, rng, calibration))

    quarter_pairs = [(0, 1), (4, 5), (2, 3), (6, 7)]
    quarters = []
    for idx, (a, b) in enumerate(quarter_pairs, start=97):
        team1 = octavos[a]
        team2 = octavos[b]
        stage_counts["quarterfinal"].update([team1, team2])
        quarters.append(simulate_knockout_match(idx, team1, team2, teams, context, rng, calibration))

    semi_pairs = [(0, 1), (2, 3)]
    semis = []
    for idx, (a, b) in enumerate(semi_pairs, start=101):
        team1 = quarters[a]
        team2 = quarters[b]
        stage_counts["semifinal"].update([team1, team2])
        semis.append(simulate_knockout_match(idx, team1, team2, teams, context, rng, calibration))

    stage_counts["final"].update(semis)
    champion = simulate_knockout_match(104, semis[0], semis[1], teams, context, rng, calibration)
    stage_counts["champion"].add(champion)
    return champion, stage_counts


def increment_counter(counter, keys):
    for key in keys:
        counter[key] = counter.get(key, 0) + 1


def top_probabilities(counter, trials, limit=12):
    rows = [
        {"team": team, "probability": round(count / trials, 4)}
        for team, count in sorted(counter.items(), key=lambda item: (-item[1], item[0]))
    ]
    return rows[:limit]


def run_monte_carlo(wb, teams, fixtures, context, calibration, trials=MONTE_CARLO_TRIALS, seed=20260610):
    rng = random.Random(seed)
    counters = {
        "group_winner": {},
        "advance_group": {},
        "round_of_32": {},
        "round_of_16": {},
        "quarterfinal": {},
        "semifinal": {},
        "final": {},
        "champion": {},
    }
    third_combinations = {}

    for _ in range(trials):
        standings, best_thirds, third_groups = simulated_group_results(fixtures, teams, context, rng, calibration)
        third_combinations[third_groups] = third_combinations.get(third_groups, 0) + 1
        advance_group = []
        for rows in standings.values():
            counters["group_winner"][rows[0]["team"]] = counters["group_winner"].get(rows[0]["team"], 0) + 1
            advance_group.extend([rows[0]["team"], rows[1]["team"]])
        advance_group.extend([row["team"] for row in best_thirds])
        increment_counter(counters["advance_group"], advance_group)

        round32 = build_round_of_32(wb, standings, best_thirds, third_groups)
        champion, stages = simulate_knockout(round32, teams, context, rng, calibration)
        for stage, teams_in_stage in stages.items():
            increment_counter(counters[stage], teams_in_stage)

    return {
        "trials": trials,
        "seed": seed,
        "champion": top_probabilities(counters["champion"], trials, limit=16),
        "final": top_probabilities(counters["final"], trials, limit=16),
        "semifinal": top_probabilities(counters["semifinal"], trials, limit=16),
        "quarterfinal": top_probabilities(counters["quarterfinal"], trials, limit=16),
        "round_of_16": top_probabilities(counters["round_of_16"], trials, limit=16),
        "advance_group": top_probabilities(counters["advance_group"], trials, limit=20),
        "group_winner": top_probabilities(counters["group_winner"], trials, limit=20),
        "third_place_combinations": top_probabilities(third_combinations, trials, limit=10),
    }


def write_workbook(wb, group_scores, knockout_results):
    ws = wb["Grupos"]
    for score in group_scores.values():
        ws[f"D{score['row']}"] = score["home_goals"]
        ws[f"E{score['row']}"] = score["away_goals"]

    ws = wb["Llaves"]
    by_row = {row: result for row, result in zip(range(3, 35), knockout_results)}
    for row, result in by_row.items():
        ws[f"D{row}"] = result["g1"]
        ws[f"E{row}"] = result["g2"]
        ws[f"G{row}"] = result["pen"]

    wb.save(OUTPUT)


def main():
    wb = openpyxl.load_workbook(INPUT, data_only=False)
    calibration = load_score_calibration()
    context = load_model_context()
    teams, team_group, fixtures = read_data(wb)
    group_scores, standings, best_thirds, third_groups, rationale = group_results(fixtures, teams, context, calibration)
    round32 = build_round_of_32(wb, standings, best_thirds, third_groups)
    ko_results = knockout(round32, teams, context, calibration)
    monte_carlo = run_monte_carlo(wb, teams, fixtures, context, calibration)
    write_workbook(wb, group_scores, ko_results)

    summary = {
        "model": {
            "base": "Sheet strength from hidden ranking/tiebreaker field",
            "adjustments": [
                "fifa_top20_rankings",
                "recent_results",
                "squad_age_caps",
                "coach_adjustment",
                "venue_climate_travel",
                "host",
                "form_expert",
                "heat_travel",
                "energy_experience",
                "upset_profile",
            ],
            "third_place_combination": third_groups,
            "monte_carlo_trials": MONTE_CARLO_TRIALS,
            "score_calibration_matches": calibration.get("matches", 0),
            "score_calibration_sources": calibration.get("sources", {}),
            "fifa_top20_teams_used": len(context.get("fifa", {})),
            "recent_form_teams_used": len(context.get("recent_form", {})),
            "squad_profile_teams_used": len(context.get("squad_profile", {})),
            "coach_adjustment_teams_used": len(context.get("coaches", {})),
            "venue_context_matches_used": len(context.get("venues", {})),
        },
        "group_standings": {g: [row["team"] for row in rows] for g, rows in standings.items()},
        "best_thirds": [row["team"] for row in best_thirds],
        "round_of_32": [{"match": m, "team1": a, "team2": b} for m, a, b in round32],
        "knockout": ko_results,
        "champion": ko_results[-1]["winner"],
        "third_place": ko_results[-2]["winner"],
        "monte_carlo": monte_carlo,
        "rationale": rationale,
    }
    SUMMARY.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"output": str(OUTPUT), "summary": str(SUMMARY), "champion": summary["champion"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
